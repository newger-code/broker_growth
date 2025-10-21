#!/usr/bin/env python3
"""
Extract commission calculations from the Excel file to understand their methodology
"""

import openpyxl
from decimal import Decimal

wb = openpyxl.load_workbook('/Users/jimnewgent/Projects/broker_growth/Docs/2025 Sept Commission Calcs.xlsx', data_only=True)
sheet = wb['ALLPayOut2']

print("=" * 120)
print("THEIR COMMISSION CALCULATIONS FROM ALLPayOut2 TAB")
print("=" * 120)

# ACQ AGENTS (Columns B-H)
print("\n" + "=" * 120)
print("ACQUISITION AGENTS")
print("=" * 120)
print(f"{'Agent':<20s} {'GP':>12s} {'Rate':>8s} {'Commission':>12s} {'Sprints':>10s} {'Total':>12s} {'Notes':<30s}")
print("-" * 120)

total_acq_comm = Decimal('0')
total_acq_sprints = Decimal('0')
total_acq_payout = Decimal('0')

row = 3
while True:
    agent = sheet.cell(row, 2).value  # Column B
    if not agent:
        break

    gp = sheet.cell(row, 3).value or 0  # Column C
    rate = sheet.cell(row, 4).value or 0  # Column D
    commission = sheet.cell(row, 5).value or 0  # Column E
    sprints = sheet.cell(row, 6).value or 0  # Column F
    total = sheet.cell(row, 7).value or 0  # Column G
    notes = sheet.cell(row, 8).value or ''  # Column H

    print(f"{str(agent):<20s} ${float(gp):>11,.2f} {float(rate)*100:>7.1f}% ${float(commission):>11,.2f} ${float(sprints):>9,.2f} ${float(total):>11,.2f} {str(notes):<30s}")

    total_acq_comm += Decimal(str(commission))
    total_acq_sprints += Decimal(str(sprints))
    total_acq_payout += Decimal(str(total))

    row += 1

print("-" * 120)
print(f"{'TOTALS':<20s} {'':>12s} {'':>8s} ${total_acq_comm:>11,.2f} ${total_acq_sprints:>9,.2f} ${total_acq_payout:>11,.2f}")

# DISPO AGENTS (Columns K-X)
print("\n\n" + "=" * 120)
print("DISPOSITION AGENTS")
print("=" * 120)
print(f"{'Agent':<20s} {'Off-Mkt GP':>12s} {'FF GP':>12s} {'Total GP':>12s} {'Tier':>5s} {'Off%':>6s} {'FF%':>7s} {'Off Pay':>12s} {'FF Pay':>12s} {'Comm':>12s} {'Sprint':>10s} {'Total':>12s}")
print("-" * 120)

total_dispo_comm = Decimal('0')
total_dispo_sprints = Decimal('0')
total_dispo_payout = Decimal('0')

row = 3
while True:
    agent = sheet.cell(row, 11).value  # Column K
    if not agent:
        break

    off_gp = sheet.cell(row, 12).value or 0  # Column L
    ff_gp = sheet.cell(row, 13).value or 0  # Column M
    total_gp = sheet.cell(row, 14).value or 0  # Column N
    tier = sheet.cell(row, 15).value or 0  # Column O
    off_rate = sheet.cell(row, 16).value or 0  # Column P
    ff_rate = sheet.cell(row, 17).value or 0  # Column Q
    off_pay = sheet.cell(row, 18).value or 0  # Column R
    ff_pay = sheet.cell(row, 19).value or 0  # Column S
    commission = sheet.cell(row, 20).value or 0  # Column T
    sprints = sheet.cell(row, 21).value or 0  # Column U
    mentor = sheet.cell(row, 22).value or 0  # Column V
    total = sheet.cell(row, 23).value or 0  # Column W
    notes = sheet.cell(row, 24).value or ''  # Column X

    print(f"{str(agent):<20s} ${float(off_gp):>11,.0f} ${float(ff_gp):>11,.0f} ${float(total_gp):>11,.0f} {str(tier):>5s} {float(off_rate)*100:>5.1f}% {float(ff_rate)*100:>6.2f}% ${float(off_pay):>11,.2f} ${float(ff_pay):>11,.2f} ${float(commission):>11,.2f} ${float(sprints):>9,.2f} ${float(total):>11,.2f}")

    total_dispo_comm += Decimal(str(commission))
    total_dispo_sprints += Decimal(str(sprints))
    total_dispo_payout += Decimal(str(total))

    row += 1

print("-" * 120)
print(f"{'TOTALS':<20s} {'':>12s} {'':>12s} {'':>12s} {'':>5s} {'':>6s} {'':>7s} {'':>12s} {'':>12s} ${total_dispo_comm:>11,.2f} ${total_dispo_sprints:>9,.2f} ${total_dispo_payout:>11,.2f}")

print("\n\n" + "=" * 120)
print("SUMMARY")
print("=" * 120)
print(f"Acq Agents Commission:     ${total_acq_comm:>15,.2f}")
print(f"Acq Agents Sprints:        ${total_acq_sprints:>15,.2f}")
print(f"Acq Agents Total:          ${total_acq_payout:>15,.2f}")
print()
print(f"Dispo Agents Commission:   ${total_dispo_comm:>15,.2f}")
print(f"Dispo Agents Sprints:      ${total_dispo_sprints:>15,.2f}")
print(f"Dispo Agents Total:        ${total_dispo_payout:>15,.2f}")
print()
print(f"AGENT SUBTOTAL:            ${total_acq_payout + total_dispo_payout:>15,.2f}")
print()
print("(Still need to add Manager and UW commissions)")
