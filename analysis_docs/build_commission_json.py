#!/usr/bin/env python3
"""Generate commission_data.json from the Excel model.

Usage:
    python analysis_docs/build_commission_json.py         Docs/2025\ Sept\ Commission\ Calcs.xlsx         analysis_docs/commission_data.json

The script reconciles September 2025 logic (reps, managers, leadership) and
emits the JSON consumed by commission_calculator.html.
"""

from __future__ import annotations

import argparse
import json
import math
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def is_flat(flag: str | None, status: str | None, exit_strategy: str | None) -> bool:
    flag_upper = flag.strip().upper() if isinstance(flag, str) else ''
    status_upper = status.strip().upper() if isinstance(status, str) else ''
    exit_upper = exit_strategy.strip().upper() if isinstance(exit_strategy, str) else ''
    if 'FF' in status_upper:
        return True
    if 'FLAT FEE' in exit_upper:
        return True
    if flag_upper in {'YES', 'Y', 'TRUE'}:
        return True
    return False


def build_dataset(workbook_path: Path) -> dict:
    wb = load_workbook(workbook_path, data_only=True)
    ws_deals = wb['Dispo Closings']

    class Deal:
        __slots__ = ('property', 'acq_agent', 'dispo_agent', 'gp', 'trx', 'flat', 'acq_manager', 'dispo_manager', 'buyer_source')

        def __init__(self, row):
            self.property = row[0]
            self.acq_agent = row[1]
            self.dispo_agent = row[3]
            self.gp = row[7] or 0
            self.trx = row[8]
            self.flat = is_flat(row[13], row[17] if len(row) > 17 else None, row[11])
            self.acq_manager = row[10]
            self.dispo_manager = row[12]
            self.buyer_source = row[9]

    deals = [Deal(row) for row in ws_deals.iter_rows(min_row=2, values_only=True) if row[0]]

    ws_sprints = wb['Sprint Winners']
    sprints = defaultdict(float)
    for row in ws_sprints.iter_rows(min_row=2, values_only=True):
        if row[1] and row[2]:
            sprints[row[1]] += row[2]

    ws_mentor = wb['HSAssigned Deals']
    mentors = defaultdict(float)
    for row in ws_mentor.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        stage = (row[1] or '').upper()
        if 'CLOSED' not in stage:
            continue
        mentor = row[7]
        amount = row[8]
        if mentor and amount:
            mentors[mentor] += amount

    ws_summary = wb['SEPTEMBER SUMMARY']
    actual_payouts = {}
    for row in ws_summary.iter_rows(min_row=3, max_row=100, values_only=True):
        name = row[2]
        payout = row[3]
        if name and isinstance(payout, (int, float)):
            actual_payouts[name] = payout

    DISPO_TIERS = [
        (25000, 0.02), (50000, 0.03), (75000, 0.04), (100000, 0.05),
        (125000, 0.06), (150000, 0.07), (175000, 0.08), (200000, 0.09),
        (math.inf, 0.10)
    ]
    ACQ_TIERS = [
        (25000, 0.04), (50000, 0.06), (100000, 0.08), (math.inf, 0.10)
    ]
    TL_TIERS = [
        (50000, 0.01), (100000, 0.02), (math.inf, 0.03)
    ]
    DISPO_MANAGER_RATES = {
        'Joe Haupt': [(0, 0.02, 0.015), (100000, 0.025, 0.02), (200000, 0.03, 0.025)],
        'Maegan Grace': [(0, 0.015, 0.01), (100000, 0.02, 0.015), (200000, 0.025, 0.02)]
    }
    manual_dispo = {'Alec Prieto': 6}
    manager_acq = {'Luis Guzman', 'Shon Yoshida', 'Devin Buford'}
    training = {'Ex Rebuilt'}

    acq_totals = defaultdict(float)
    dispo_totals = defaultdict(lambda: {'off': 0.0, 'ff': 0.0})
    for deal in deals:
        if deal.acq_agent:
            acq_totals[deal.acq_agent] += deal.gp
        if deal.dispo_agent:
            bucket = dispo_totals[deal.dispo_agent]
            if deal.flat:
                bucket['ff'] += deal.gp
            else:
                bucket['off'] += deal.gp

    acq_agents = []
    for agent, gp in acq_totals.items():
        rate = next(rate for limit, rate in ACQ_TIERS if gp < limit)
        commission = gp * rate
        payout = actual_payouts.get(agent, 0)
        base = commission + sprints.get(agent, 0)
        if agent in manager_acq:
            role = 'manager'
            bucket = 'acq_manager_personal'
        elif agent in training:
            role = 'training'
            bucket = 'acq_training'
        else:
            role = 'rep'
            bucket = 'acq_rep'
        acq_agents.append({
            'name': agent,
            'gross_profit': gp,
            'rate': rate,
            'commission': commission,
            'sprint': sprints.get(agent, 0),
            'role': role,
            'summary_bucket': bucket,
            'actual_payout': payout,
            'manual_adjustment': payout - base
        })
    for name in ('Terrell Johnson', 'Steven Stack', 'Jarod Weaver', 'Scott Pennebaker'):
        if name in actual_payouts and name not in acq_totals:
            payout = actual_payouts[name]
            base = sprints.get(name, 0)
            acq_agents.append({
                'name': name,
                'gross_profit': 0.0,
                'rate': ACQ_TIERS[0][1],
                'commission': 0.0,
                'sprint': sprints.get(name, 0),
                'role': 'rep',
                'summary_bucket': 'acq_rep',
                'actual_payout': payout,
                'manual_adjustment': payout - base
            })

    dispo_agents = []
    for agent, vals in dispo_totals.items():
        total = vals['off'] + vals['ff']
        tier = manual_dispo.get(agent)
        if not tier:
            tier = next(idx for idx, (limit, _) in enumerate(DISPO_TIERS, start=1) if total < limit)
        rate = DISPO_TIERS[tier-1][1]
        ff_rate = rate * (2/3)
        commission = vals['off'] * rate + vals['ff'] * ff_rate
        payout = actual_payouts.get(agent, 0)
        base = commission + sprints.get(agent, 0) + mentors.get(agent, 0)
        bucket = 'dispo_manager_personal' if agent in DISPO_MANAGER_RATES else 'dispo_rep'
        dispo_agents.append({
            'name': agent,
            'off_market_gp': vals['off'],
            'flat_fee_gp': vals['ff'],
            'tier': tier,
            'off_rate': rate,
            'ff_rate': ff_rate,
            'commission': commission,
            'sprint': sprints.get(agent, 0),
            'mentor': mentors.get(agent, 0),
            'role': 'manager' if bucket == 'dispo_manager_personal' else 'rep',
            'summary_bucket': bucket,
            'actual_payout': payout,
            'manual_adjustment': payout - base
        })
    for name in ('Kayla Watkins', 'Kirk Schaafsma'):
        if name in actual_payouts and name not in dispo_totals:
            payout = actual_payouts[name]
            base = sprints.get(name, 0) + mentors.get(name, 0)
            dispo_agents.append({
                'name': name,
                'off_market_gp': 0.0,
                'flat_fee_gp': 0.0,
                'tier': 1,
                'off_rate': DISPO_TIERS[0][1],
                'ff_rate': DISPO_TIERS[0][1] * (2/3),
                'commission': 0.0,
                'sprint': sprints.get(name, 0),
                'mentor': mentors.get(name, 0),
                'role': 'rep',
                'summary_bucket': 'dispo_rep',
                'actual_payout': payout,
                'manual_adjustment': payout - base
            })

    managers = []
    for manager in manager_acq:
        gp = acq_totals.get(manager, 0)
        rate = next(rate for limit, rate in ACQ_TIERS if gp < limit)
        personal_comm = gp * rate
        team_gp = sum(deal.gp for deal in deals if deal.acq_manager == manager)
        team_rate = next(rate for limit, rate in TL_TIERS if team_gp < limit)
        team_comm = team_gp * team_rate
        base = personal_comm + team_comm + sprints.get(manager, 0)
        payout = actual_payouts.get(manager, 0)
        managers.append({
            'name': manager,
            'type': 'acquisition_manager',
            'summary_bucket': 'acq_manager',
            'personal_gp': gp,
            'personal_rate': rate,
            'personal_commission': personal_comm,
            'team_gp': team_gp,
            'team_rate': team_rate,
            'team_commission': team_comm,
            'sprint': sprints.get(manager, 0),
            'actual_payout': payout,
            'manual_adjustment': payout - base
        })

    for manager, tiers in DISPO_MANAGER_RATES.items():
        personal = next(d for d in dispo_agents if d['name'] == manager)
        team_off = sum(deal.gp for deal in deals if deal.dispo_manager == manager and not deal.flat)
        team_ff = sum(deal.gp for deal in deals if deal.dispo_manager == manager and deal.flat)
        team_total = team_off + team_ff
        off_rate, ff_rate = tiers[0][1], tiers[0][2]
        for threshold, o_rate, f_rate in tiers:
            if team_total >= threshold:
                off_rate, ff_rate = o_rate, f_rate
        team_comm = team_off * off_rate + team_ff * ff_rate
        base = personal['commission'] + team_comm + personal['sprint'] + personal['mentor']
        payout = actual_payouts.get(manager, 0)
        managers.append({
            'name': manager,
            'type': 'disposition_manager',
            'summary_bucket': 'dispo_manager',
            'personal_off_gp': personal['off_market_gp'],
            'personal_ff_gp': personal['flat_fee_gp'],
            'personal_commission': personal['commission'],
            'team_off_gp': team_off,
            'team_ff_gp': team_ff,
            'team_off_rate': off_rate,
            'team_ff_rate': ff_rate,
            'team_commission': team_comm,
            'sprint': personal['sprint'],
            'mentor': personal['mentor'],
            'actual_payout': payout,
            'manual_adjustment': payout - base
        })

    def extract(sheet_name: str):
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=6, max_row=20):
            if row[1].value == 'September':
                return {i: cell.value for i, cell in enumerate(row)}
        raise RuntimeError(f'Month row not found in {sheet_name}')

    sol = extract('Solomon')
    comp_component = (sol[8] or 0) + (sol[9] or 0)
    team_component = (sol[19] or 0) + (sol[20] or 0)
    weights = {'company': sol[22] or 0, 'team': sol[23] or 0}
    base = comp_component * weights['company'] + team_component * weights['team']
    payout = actual_payouts.get('Patrick Solomon', sol[24])
    managers.append({
        'name': 'Patrick Solomon',
        'type': 'senior_leader',
        'summary_bucket': 'senior_leader',
        'company_gp': sol[2],
        'company_trx': sol[3],
        'company_rates': {'gp': sol[5], 'trx': sol[6]},
        'team_gp': sol[13],
        'team_trx': sol[14],
        'team_rates': {'gp': sol[16], 'trx': sol[17]},
        'weights': weights,
        'actual_payout': payout,
        'manual_adjustment': payout - base
    })

    gor = extract('Gorski')
    comp_component = (gor[8] or 0) + (gor[9] or 0)
    team_component = (gor[19] or 0) + (gor[20] or 0)
    weights = {'company': gor[22] or 0, 'team': gor[23] or 0}
    base = comp_component * weights['company'] + team_component * weights['team']
    payout = actual_payouts.get('Rob Gorski', (gor[24] or 0) + (gor[25] or 0))
    managers.append({
        'name': 'Rob Gorski',
        'type': 'senior_leader',
        'summary_bucket': 'senior_leader',
        'company_gp': gor[2],
        'company_trx': gor[3],
        'company_rates': {'gp': gor[5], 'trx': gor[6]},
        'team_gp': gor[13],
        'team_trx': gor[14],
        'team_rates': {'gp': gor[16], 'trx': gor[17]},
        'weights': weights,
        'actual_payout': payout,
        'manual_adjustment': payout - base
    })

    payout_uw = actual_payouts.get('Dustin Hepburn')
    if payout_uw is None:
        hep = wb['Hepburn']
        for row in hep.iter_rows(min_row=24, max_row=40):
            if row[1].value == 'Total Bonus':
                payout_uw = row[2].value
                break
    managers.append({
        'name': 'Dustin Hepburn',
        'type': 'underwriting',
        'summary_bucket': 'underwriting',
        'actual_payout': payout_uw,
        'manual_adjustment': 0
    })

    return {
        'metadata': {
            'source_file': str(workbook_path.name)
        },
        'acquisition_agents': acq_agents,
        'disposition_agents': dispo_agents,
        'managers': managers
    }


def main() -> None:
    parser = argparse.ArgumentParser(description='Generate commission_data.json from the Excel model.')
    parser.add_argument('workbook', type=Path, help='Path to the Commission Calcs workbook (.xlsx)')
    parser.add_argument('output', type=Path, help='Path to write commission_data.json')
    args = parser.parse_args()

    dataset = build_dataset(args.workbook)
    args.output.write_text(json.dumps(dataset, indent=2))
    print(f'Wrote {args.output}')


if __name__ == '__main__':
    main()
