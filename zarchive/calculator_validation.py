#!/usr/bin/env python3
"""
Calculate what the commission calculator SHOULD produce based on its code logic
Compare to Excel SEPTEMBER SUMMARY
"""

from openpyxl import load_workbook

# Tier functions from calculator
def get_dispo_tier(gp):
    if gp < 25000: return {'tier': 1, 'off': 0.02, 'ff': 0.013333}
    if gp < 50000: return {'tier': 2, 'off': 0.03, 'ff': 0.02}
    if gp < 75000: return {'tier': 3, 'off': 0.04, 'ff': 0.026667}
    if gp < 100000: return {'tier': 4, 'off': 0.05, 'ff': 0.033333}
    if gp < 125000: return {'tier': 5, 'off': 0.06, 'ff': 0.04}
    if gp < 150000: return {'tier': 6, 'off': 0.07, 'ff': 0.046667}
    if gp < 175000: return {'tier': 7, 'off': 0.08, 'ff': 0.053333}
    if gp < 200000: return {'tier': 8, 'off': 0.09, 'ff': 0.06}
    return {'tier': 9, 'off': 0.10, 'ff': 0.066667}

def get_acq_tier(gp):
    if gp < 25000: return {'tier': 1, 'rate': 0.04}
    if gp < 50000: return {'tier': 2, 'rate': 0.06}
    if gp < 100000: return {'tier': 3, 'rate': 0.08}
    return {'tier': 4, 'rate': 0.10}

# Sprint data from calculator lines 995-1018
sprints = {
    'Brittany Taylor': 500, 'Terrell Johnson': 400, 'Alec Prieto': 350,
    'Kyle Singer': 550, 'Ian Ross': 300, 'Jesse Sychowski': 300,
    'Joe Haupt': 300, 'Garrett Paschal': 200, 'Steven Stack': 200,
    'Ashley Preston': 150, 'Luis Guzman': 150, 'Christian Flasch': 100,
    'Devin Buford': 100, 'Devin Cooper': 100, 'Devin Hoffman': 100,
    'Jarod Weaver': 100, 'Kayla Watkins': 100, 'Kirk Schaafsma': 100,
    'Mel Grant': 100, 'Miguel Aguilar': 100, 'Tamara Humbolt': 100,
    'Warren Smith': 100
}

# DISPO data from calculator lines 1022-1117
dispo = {
    'Alec Prieto': {'off': 93403, 'ff': 31569, 'total': 124972},
    'Brittany Taylor': {'off': 17867, 'ff': 7416, 'total': 25283},
    'Christian Flasch': {'off': 16582, 'ff': 37714, 'total': 54296},
    'Devin Cooper': {'off': 33725, 'ff': 27760, 'total': 61485},
    'Devin Hoffman': {'off': 0, 'ff': 3235, 'total': 3235},
    'Ex Rebuilt': {'off': 10000, 'ff': 6500, 'total': 16500},
    'Jack Webster': {'off': 0, 'ff': 41863, 'total': 41863},
    'Joe Haupt': {'off': 58190, 'ff': 59133, 'total': 117323},
    'Leland Boyd': {'off': 37244, 'ff': 0, 'total': 37244},
    'Maegan Grace': {'off': 37750, 'ff': 21515, 'total': 59265},
    'Mel Grant': {'off': 4000, 'ff': 0, 'total': 4000},
    'Miguel Aguilar': {'off': 46069, 'ff': 0, 'total': 46069},
    'Tamara Humbolt': {'off': 15428, 'ff': 15353, 'total': 30781},
    'Vincent Gnapi': {'off': 46503, 'ff': 4629, 'total': 51132},
    'Yoseph Israel': {'off': 11600, 'ff': 0, 'total': 11600}
}

# ACQ data from calculator lines 1121-1222
acq = {
    'Andrew Caceres': 77053,
    'Ashley Preston': 60123,
    'Chris Chambers': 29225,
    'Devin Buford': 29005,
    'Dominick Mazliah': 77787,
    'Ex Rebuilt': 76443,
    'Garrett Paschal': 3235,
    'Ian Ross': 3000,
    'Jesse Sychowski': 61375,
    'Kyle Singer': 83227,
    'Luis Guzman': 60303,
    'Scott Pennebaker': 10000,
    'Shon Yoshida': 71552,
    'Steve Shelburne': 6000,
    'Warren Smith': 33715,
    'Terrell Johnson': 0,
    'Steven Stack': 0,
    'Jarod Weaver': 0,
    'Kayla Watkins': 0,
    'Kirk Schaafsma': 0
}

# Load Excel
wb = load_workbook('/Users/jimnewgent/Projects/broker_growth/Docs/2025 Sept Commission Calcs.xlsx', data_only=True)
sheet = wb['SEPTEMBER SUMMARY']

excel_data = {}
for row in range(3, 39):
    name = sheet.cell(row=row, column=3).value
    payout = sheet.cell(row=row, column=4).value
    if name and payout:
        excel_data[name] = payout

print("=" * 100)
print("CALCULATOR vs EXCEL VALIDATION")
print("=" * 100)

print("\nDISPO AGENTS:")
print(f"{'Name':<20} {'Off GP':<12} {'FF GP':<12} {'Total GP':<12} {'Tier':<6} {'Calc Comm':<12} {'Sprint':<8} {'Calc Total':<12} {'Excel':<12} {'Diff':<10}")
print("-" * 100)

# Exclusions (from calculator lines 1299-1301)
manager_dispo = ['Joe Haupt', 'Maegan Grace']
manager_acq = ['Luis Guzman', 'Shon Yoshida', 'Devin Buford']
exclude_acq = ['Ex Rebuilt', 'Scott Pennebaker']

dispo_calc_total = 0
dispo_excel_total = 0

for name, data in sorted(dispo.items()):
    if name in manager_dispo:
        continue  # Skip managers
    off_gp = data['off']
    ff_gp = data['ff']
    total_gp = data['total']

    # Special cases
    if name == 'Alec Prieto':
        # Alec gets natural tier + 1 = Tier 5 + 1 = Tier 6
        tier = {'tier': 6, 'off': 0.07, 'ff': 0.046667}
    elif name == 'Ex Rebuilt':
        tier = get_dispo_tier(0)  # Always tier 1
    else:
        tier = get_dispo_tier(total_gp)

    base_comm = off_gp * tier['off'] + ff_gp * tier['ff']
    sprint = sprints.get(name, 0)
    calc_total = base_comm + sprint
    excel_total = excel_data.get(name, 0)
    diff = calc_total - excel_total

    print(f"{name:<20} ${off_gp:<11,} ${ff_gp:<11,} ${total_gp:<11,} {tier['tier']:<6} ${base_comm:<11.2f} ${sprint:<7} ${calc_total:<11.2f} ${excel_total:<11.2f} ${diff:>9.2f}")

    dispo_calc_total += calc_total
    dispo_excel_total += excel_total

print("\nACQ AGENTS:")
print(f"{'Name':<20} {'GP':<12} {'Tier':<6} {'Calc Comm':<12} {'Sprint':<8} {'Calc Total':<12} {'Excel':<12} {'Diff':<10}")
print("-" * 100)

acq_calc_total = 0
acq_excel_total = 0

for name, gp in sorted(acq.items()):
    if name in manager_acq or name in exclude_acq:
        continue  # Skip managers and excluded agents
    if gp == 0:  # Sprint-only
        calc_total = sprints.get(name, 0)
        excel_total = excel_data.get(name, 0)
        diff = calc_total - excel_total
        print(f"{name:<20} ${gp:<11,} {'N/A':<6} ${0.00:<11.2f} ${sprints.get(name, 0):<7} ${calc_total:<11.2f} ${excel_total:<11.2f} ${diff:>9.2f}")
    else:
        tier = get_acq_tier(gp) if name != 'Ex Rebuilt' else {'tier': 1, 'rate': 0.04}
        base_comm = gp * tier['rate']
        sprint = sprints.get(name, 0)
        calc_total = base_comm + sprint
        excel_total = excel_data.get(name, 0)
        diff = calc_total - excel_total
        print(f"{name:<20} ${gp:<11,} {tier['tier']:<6} ${base_comm:<11.2f} ${sprint:<7} ${calc_total:<11.2f} ${excel_total:<11.2f} ${diff:>9.2f}")

    acq_calc_total += calc_total
    acq_excel_total += excel_total

print("\n" + "=" * 100)
print(f"{'DISPO SUBTOTAL':<20} ${dispo_calc_total:>11.2f} vs Excel ${dispo_excel_total:>11.2f} | Diff: ${dispo_calc_total - dispo_excel_total:>9.2f}")
print(f"{'ACQ SUBTOTAL':<20} ${acq_calc_total:>11.2f} vs Excel ${acq_excel_total:>11.2f} | Diff: ${acq_calc_total - acq_excel_total:>9.2f}")
print(f"{'AGENTS TOTAL':<20} ${dispo_calc_total + acq_calc_total:>11.2f} vs Excel ${dispo_excel_total + acq_excel_total:>11.2f} | Diff: ${(dispo_calc_total + acq_calc_total) - (dispo_excel_total + acq_excel_total):>9.2f}")

# Managers (from Excel)
managers_excel = {
    'Luis Guzman': 8780.43,
    'Shon Yoshida': 7513.40,
    'Patrick Solomon': 6043.56,
    'Devin Buford': 1900.30,
    'Joe Haupt': 8452.81,
    'Rob Gorski': 3672.31,
    'Maegan Grace': 3260.74,
    'Dustin Hepburn': 1346.26
}

manager_total = sum(managers_excel.values())
print(f"{'MANAGERS TOTAL':<20} ${manager_total:>11.2f} (from Excel)")

grand_calc = dispo_calc_total + acq_calc_total + manager_total
grand_excel = 97539.31
print(f"\n{'GRAND TOTAL':<20} ${grand_calc:>11.2f} vs Excel ${grand_excel:>11.2f} | Diff: ${grand_calc - grand_excel:>9.2f}")
